"""Supplier price list import — CSV/XLSX parsing, validation, and upsert preview."""
from __future__ import annotations

import csv
import io
from typing import Any

from sqlalchemy.orm import Session

from app import models, schemas
from app.config import get_settings
from app.services.pricing import price_from_markup
from app.services.decision_service import evaluate_product_decision

settings = get_settings()

IMPORT_COLUMNS = {
    "sku",
    "name",
    "description",
    "category",
    "brand",
    "cost_price",
    "stock_quantity",
    "currency",
    "image_url",
}


def _normalize_row(raw: dict[str, Any]) -> dict[str, str]:
    return {str(k).strip().lower(): ("" if v is None else str(v).strip()) for k, v in raw.items() if k}


def parse_csv(content: bytes) -> list[dict[str, str]]:
    text = content.decode("utf-8-sig")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    return [_normalize_row(row) for row in reader]


def parse_xlsx(content: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
    parsed: list[dict[str, str]] = []
    for values in rows_iter:
        if not values or all(v is None or str(v).strip() == "" for v in values):
            continue
        row: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            val = values[idx] if idx < len(values) else None
            row[header] = "" if val is None else str(val).strip()
        parsed.append(row)
    wb.close()
    return parsed


def parse_upload(filename: str, content: bytes) -> list[dict[str, str]]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return parse_csv(content)
    if name.endswith(".xlsx"):
        return parse_xlsx(content)
    raise ValueError(
        "Неподдерживаемый формат файла. Сейчас поддерживаются только CSV и Excel (.xlsx). "
        "PDF, Word и изображения будут добавлены позже."
    )


def _find_existing(db: Session, supplier_id: str, sku: str) -> models.Product | None:
    return (
        db.query(models.Product)
        .filter(models.Product.supplier_id == supplier_id, models.Product.sku == sku)
        .first()
    )


def build_preview_rows(
    raw_rows: list[dict[str, str]],
    supplier_id: str,
    db: Session,
) -> list[schemas.ProductImportPreviewRow]:
    seen_skus: set[str] = set()
    rows: list[schemas.ProductImportPreviewRow] = []

    for raw in raw_rows:
        sku = (raw.get("sku") or "").strip()
        name = (raw.get("name") or "").strip()
        error: str | None = None
        row_status: str = "error"

        if not sku:
            error = "Поле 'sku' обязательно"
        elif sku in seen_skus:
            error = "Дублирующий SKU в файле"
        elif not name:
            error = "Поле 'name' обязательно"
        else:
            try:
                cost_price = float(raw.get("cost_price") or 0)
            except ValueError:
                cost_price = 0.0
                error = "Некорректная цена закупки"
            else:
                if cost_price <= 0:
                    error = "Некорректная цена закупки"

        if error is None:
            seen_skus.add(sku)
            existing = _find_existing(db, supplier_id, sku)
            row_status = "update" if existing else "new"
            cost_price = float(raw.get("cost_price") or 0)
        else:
            cost_price = 0.0

        try:
            stock_quantity = int(float(raw.get("stock_quantity") or 0))
        except ValueError:
            stock_quantity = 0

        markup = settings.DEFAULT_MARKUP_PERCENT
        suggested_price = price_from_markup(cost_price, markup) if cost_price > 0 else 0.0

        rows.append(
            schemas.ProductImportPreviewRow(
                sku=sku or None,
                name=name or "(без названия)",
                description=raw.get("description") or None,
                category=raw.get("category") or None,
                brand=raw.get("brand") or None,
                cost_price=cost_price,
                stock_quantity=stock_quantity,
                currency=raw.get("currency") or settings.DEFAULT_CURRENCY,
                image_url=raw.get("image_url") or None,
                suggested_selling_price=suggested_price,
                row_status=row_status,
                valid=row_status in ("new", "update"),
                error=error,
            )
        )

    return rows


def commit_import(
    supplier_id: str,
    rows: list[schemas.ProductImportPreviewRow],
    db: Session,
) -> schemas.ProductImportCommitResponse:
    supplier = db.query(models.Supplier).filter(models.Supplier.id == supplier_id).first()
    if not supplier:
        raise ValueError("Supplier not found")

    added_count = 0
    updated_count = 0
    skipped_count = 0
    error_count = 0
    product_ids: list[str] = []
    details: list[schemas.ProductImportCommitDetailRow] = []
    markup = settings.DEFAULT_MARKUP_PERCENT

    for row in rows:
        if row.row_status == "error" or not row.valid:
            error_count += 1
            details.append(
                schemas.ProductImportCommitDetailRow(
                    sku=row.sku,
                    name=row.name,
                    action="error",
                    error=row.error or "Строка с ошибкой",
                )
            )
            continue

        if not row.sku:
            skipped_count += 1
            details.append(
                schemas.ProductImportCommitDetailRow(
                    sku=row.sku,
                    name=row.name,
                    action="skipped",
                    error="Пропущено: пустой SKU",
                )
            )
            continue

        selling_price = price_from_markup(row.cost_price, markup)
        existing = _find_existing(db, supplier.id, row.sku)

        if existing:
            existing.name_raw = row.name
            existing.description_raw = row.description
            existing.category = row.category
            existing.brand = row.brand
            existing.cost_price = row.cost_price
            existing.stock_quantity = row.stock_quantity
            existing.currency = row.currency
            existing.image_url = row.image_url
            existing.selling_price = selling_price
            existing.markup_percent = markup
            db.flush()
            updated_count += 1
            product_ids.append(existing.id)
            details.append(
                schemas.ProductImportCommitDetailRow(
                    sku=row.sku,
                    name=row.name,
                    action="updated",
                    product_id=existing.id,
                )
            )
        else:
            product = models.Product(
                supplier_id=supplier.id,
                sku=row.sku,
                name_raw=row.name,
                description_raw=row.description,
                category=row.category,
                brand=row.brand,
                cost_price=row.cost_price,
                selling_price=selling_price,
                markup_percent=markup,
                stock_quantity=row.stock_quantity,
                currency=row.currency,
                status="draft",
                image_url=row.image_url,
            )
            db.add(product)
            db.flush()
            added_count += 1
            product_ids.append(product.id)
            details.append(
                schemas.ProductImportCommitDetailRow(
                    sku=row.sku,
                    name=row.name,
                    action="added",
                    product_id=product.id,
                )
            )

    db.commit()

    good_count = risk_count = bad_count = 0
    for pid in product_ids:
        product = db.query(models.Product).filter(models.Product.id == pid).first()
        if not product:
            continue
        status, _, _ = evaluate_product_decision(
            product.cost_price,
            product.selling_price,
            product.stock_quantity or 0,
            product.markup_percent,
        )
        if status == "good":
            good_count += 1
        elif status == "risk":
            risk_count += 1
        else:
            bad_count += 1

    return schemas.ProductImportCommitResponse(
        added_count=added_count,
        updated_count=updated_count,
        skipped_count=skipped_count,
        error_count=error_count,
        good_count=good_count,
        risk_count=risk_count,
        bad_count=bad_count,
        product_ids=product_ids,
        rows=details,
    )
